#coding:utf-8

from openpyxl import Workbook
from openpyxl import load_workbook

#创建一个工作簿
wb = Workbook()

# 获取活动工作表
ws = wb.active

# 将数据写入单元格
ws['A1'] = 42

# 将list以追加写入行中
ws.append([1, 2, 3])

# 在execl中插入一张新的工作表(最后位置)
ws1 = wb.create_sheet("Mysheet")

# 在execl中插入一张新的工作表(插入第一位置)
ws2 = wb.create_sheet("Mysheet", 0)

# Python类型自动转换   将A3数据填充为当前时间
import datetime
ws['A3'] = datetime.datetime.now()

# 保存该文件
wb.save("sample.xlsx")