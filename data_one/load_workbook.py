#coding:utf-8

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, Fill
from openpyxl.styles import colors
from openpyxl.styles import Fill,fills
from openpyxl.formatting.rule import ColorScaleRule

#打开EXECL文件
# wbk = load_workbook('python_openxl.xlsx')

# 加载已存在的execl的文件
filepath = "E:\python\data\data_one\python_openxl.xlsx"
wb = load_workbook(filename=filepath)

#获取每个sheet名称
sheetnames = wb.sheetnames

#获取第一个sheet
ws = wb.sheetnames[0]

# print sheetnames

# 遍历execl中的工作表名
for sheet in wb:
    print sheet.title

# 通过索引加载sheet index从0开始
wn = wb['jobs']

# 修改工作表名称
# wn.title = 'jobs'

#
wn['A4'] = 'hello,world'

# # 文本对齐方式
# align = Alignment(horizontal='center', vertical='center')
# ws.cell(row=deng_lu_taskRow, column=index + 3).alignment = align
#
# # 字体大小
# font = Font(size=10)
# ws.cell(row=taskRow, column=column).font = font

for i in xrange(0,10000):
    pass


# 保存文件
wb.save(filepath)



